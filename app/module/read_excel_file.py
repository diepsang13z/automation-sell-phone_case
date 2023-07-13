import openpyxl
import re


def getProductColumn(excel_file):
    """Returns the product column content in the excel file"""
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    n_row = ws.max_row
    products = []
    for i in range(2, n_row + 1):
        cell = ws.cell(row=i, column=3).value
        products.append(cell)
    return products


def getProductDict(products_column):
    """Returns a product dictionary with structure: {name : {code : quantity}}"""
    products_dict = {}
    pattern_product_name = r"Tên phân loại hàng:([\w -/]+),([\w -/]+);"
    pattern_product_quantity = r"Số lượng: (\d+);"

    for products_cell in products_column:
        lines = products_cell.split("\n")

        for line in lines:
            match = re.search(pattern_product_name, line)

            if match == None:
                continue

            phone_name = match.group(2).replace("/", "-")
            product_code = match.group(1)
            product_quantity = re.search(pattern_product_quantity, line).group(1)

            products_dict.setdefault(phone_name, {})
            if products_dict[phone_name].get(product_code) == None:
                products_dict[phone_name][product_code] = product_quantity
            else:
                old_quantity = products_dict[phone_name][product_code]
                new_quantity = int(old_quantity) + int(product_quantity)
                products_dict[phone_name][product_code] = new_quantity

    return products_dict
